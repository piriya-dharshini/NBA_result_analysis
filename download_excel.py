from openpyxl import Workbook
import os
import mysql.connector
mydb = mysql.connector.connect(
   host =  "localhost",
   user = "root",
   password = "pdk164@#",
   database = "Result_Analysis"
)

cur = mydb.cursor()

def write_to_excel(filename, column_name1, column_name2, data_rows, column_names_grades1, grade_data, subsum_header, subsum_stat, oa_summary):

    # print(column_names_grades1, column_names_grades2, grade_data)
    filename = filename + '.xlsx'
    wb = Workbook()
    # Select the active worksheet (by default, it's the first sheet)
    ws = wb.active
    # Rename the worksheet to 'result_data'
    ws.title = 'Result Data'

    # Write column names to the first row
    for col_idx, column_name in enumerate(column_name1, start = 1):
        ws.cell(row = 1, column = col_idx, value = column_name)
    for col_idx, column_name in enumerate(column_name2, start = 1):
        ws.cell(row = 2, column = col_idx, value = column_name)

    # Write data rows starting from the second row
    for row_idx, row in enumerate(data_rows, start = 3):
        for col_idx, cell_value in enumerate(row, start = 1):
            ws.cell(row = row_idx, column = col_idx, value = cell_value)

    ws = wb.create_sheet(title = 'Grade Analysis')#malavika,madhu,padma use this for creating sheets in the excel sheet result_data
    # Select the active worksheet (by default, it's the first sheet)
    wb.active = wb['Grade Analysis']
    # Rename the worksheet to 'result_data'
    ws.title = 'Grade Analysis'

    # Write column names to the first row
    for col_idx, column_name in enumerate(column_names_grades1, start = 1):
        ws.cell(row = 1, column = col_idx, value = column_name)

    # Write data rows starting from the second row
    for row_idx, row in enumerate(grade_data, start = 3):
        for col_idx, cell_value in enumerate(row, start = 1):
            ws.cell(row = row_idx, column = col_idx, value = cell_value)

    wb.save(filename)

    # subject wise summary
    ws = wb.create_sheet(title = 'Subject Wise Summary')

    # Select the active worksheet (by default, it's the first sheet)
    wb.active = wb['Subject Wise Summary']
    # Rename the worksheet to 'result_data'
    ws.title = 'Subject Wise Summary'

    # Write column names to the first row
    for col_idx, column_name in enumerate(subsum_header, start = 1):
        ws.cell(row = 1, column = col_idx, value = column_name)

    # Write data rows starting from the second row
    for row_idx, row in enumerate(subsum_stat, start = 3):
        for col_idx, cell_value in enumerate(row, start = 1):
            ws.cell(row = row_idx, column = col_idx, value = cell_value)

    wb.save(filename)

    # overall summary
    ws = wb.create_sheet(title = 'Overall Summary')

    # Select the active worksheet (by default, it's the first sheet)
    wb.active = wb['Overall Summary']
    # Rename the worksheet to 'result_data'
    wb.title = 'Overall Summary'

    # Write column names to the first row
    for col_idx1, column_name1 in enumerate(oa_summary[0], start = 1):
        ws.cell(row = 1, column = col_idx1, value = column_name1)

    # Write data rows starting from the third row
    for row_idx1, row1 in enumerate(oa_summary[1], start = 3):
        for col_idx1, cell_value1 in enumerate(row1, start = 1):
            ws.cell(row = row_idx1, column = col_idx1, value = cell_value1)

    # Write column names to the tenth row
    for col_idx2, column_name2 in enumerate(oa_summary[2], start = 1):
        ws.cell(row = 10, column = col_idx2, value = column_name2)

    # Write data rows starting from the twelveth row
    for row_idx2, row2 in enumerate(oa_summary[3], start = 12):
        for col_idx2, cell_value2 in enumerate(row2, start = 1):
            ws.cell(row = row_idx2, column = col_idx2, value = cell_value2)

    # Write column names to the sixteenth row
    for col_idx2, column_name2 in enumerate(oa_summary[4], start = 1):
        ws.cell(row = 16, column = col_idx2, value = column_name2)

    # Write data rows starting from the eightteenth row
    for row_idx3, row3 in enumerate(oa_summary[5], start = 18):
        for col_idx3, cell_value3 in enumerate(row3, start = 1):
            ws.cell(row = row_idx3, column = col_idx3, value = cell_value3)

    # Write column names to the twenty-fifth row
    for col_idx4, column_name4 in enumerate(oa_summary[6], start = 1):
        ws.cell(row = 25, column = col_idx4, value = column_name4)

    # Write data rows starting from the twenty-seventh row
    for row_idx4, row4 in enumerate(oa_summary[7], start = 27):
        for col_idx4, cell_value4 in enumerate(row4, start = 1):
            ws.cell(row = row_idx4, column = col_idx4, value = cell_value4)

    wb.save(filename)