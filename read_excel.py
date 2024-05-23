# NECESSARY IMPORTS OF FUCTIONS OR MODULES

import pandas as pd
import openpyxl

# USER DEFINED FUNCTIONS

def read_sheet(name):

    newworkbook = openpyxl.load_workbook(name)
    sheetnames = newworkbook.sheetnames
    # print(sheetnames[0])
    # Reading the .xlsx (EXCEL file) given the filename and sheetname
    all_sheets = pd.read_excel(name, sheetnames[0])

    # Replaces nan with a dashes wherever nan is found in the dataframe
    all_sheets = all_sheets.fillna('-')

    list_of_column_names = []
    list_of_column_values = []

    # Adding the column / field names to the list - list_of_column_names (LIST)
    for key, value in all_sheets.items():
        list_of_column_names.append(key)

    # Adding the values present in the column to the list - list_of_column_values (NESTED LIST)
    for item in list(all_sheets):
        list_of_column_values.append(all_sheets[item].tolist())

    return [list_of_column_names, list_of_column_values]


def filter_data(data):

    # Unpacking the return value of the function read_sheet()
    column_names, column_values = data

    # Removing the last few rows that analyse the result
    required_length = len([x for x in column_values[0] if isinstance(x, int) or isinstance(x, float)])
    for index in range(len(column_values)):
        column_values[index] = column_values[index][ : required_length]
    # Converting Register Number to integer type
    column_values[0] = [int(x) for x in column_values[0]]
    return [column_names, column_values]

# print(filter_data())